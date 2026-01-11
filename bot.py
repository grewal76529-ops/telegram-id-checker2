import os
import telebot
from telebot import types
from openpyxl import load_workbook
import re, time

# ================= TOKEN (KOYEB / VPS SAFE) =================
TOKEN = os.getenv("BOT_TOKEN")
if not TOKEN:
    raise ValueError("BOT_TOKEN NOT FOUND IN ENVIRONMENT VARIABLES")

OWNER_ID = 8486481425
bot = telebot.TeleBot(TOKEN, threaded=True)

# ================= CHANNELS =================
CHANNELS = ["@grewalws", "@VAXSYWORKS", "@grewalsws"]
CHANNEL_LINKS = [
    ("WORK BY AMAR", "https://t.me/grewalws"),
    ("WORK BY VAXXY", "https://t.me/VAXSYWORKS"),
    ("CHAT GC", "https://t.me/grewalsws")
]

# ================= STORAGE =================
session = {}
utility = {}
thanks_cooldown = {}
all_users = set()

# ================= PRIVATE ONLY =================
def private_only(func):
    def wrapper(msg, *a, **k):
        if msg.chat.type != "private":
            return
        return func(msg, *a, **k)
    return wrapper

# ================= FORCE JOIN =================
def is_joined(uid):
    for ch in CHANNELS:
        try:
            st = bot.get_chat_member(ch, uid).status
            if st not in ["member", "administrator", "creator"]:
                return False
        except:
            return False
    return True

def join_msg(uid):
    kb = types.InlineKeyboardMarkup()
    for n, l in CHANNEL_LINKS:
        kb.add(types.InlineKeyboardButton(n, url=l))
    kb.add(types.InlineKeyboardButton("✅ CONTINUE", callback_data="continue"))
    bot.send_message(
        uid,
        "🚫 ACCESS DENIED\n\nPLEASE JOIN ALL CHANNELS TO USE THIS BOT",
        reply_markup=kb
    )

# ================= THANK YOU MESSAGE =================
def send_thanks(uid):
    now = time.time()
    last = thanks_cooldown.get(uid, 0)
    if now - last >= 600:
        thanks_cooldown[uid] = now
        bot.send_message(
            uid,
            "🙏 THANKS FOR USING THIS BOT BY #TEAMGREWAL\n\n"
            "💬 ANY COMPLAINTS / SUGGESTIONS\n"
            "CONTACT ADMIN 👉 @grewal_65"
        )

# ================= START =================
@bot.message_handler(commands=["start"])
@private_only
def start(msg):
    uid = msg.chat.id
    all_users.add(uid)

    if not is_joined(uid):
        join_msg(uid)
        return

    session[uid] = {"state": "WAIT_EXCEL"}
    bot.send_message(
        uid,
        "✨ WELCOME TO ID VERIFICATION BOT ✨\n\n"
        "📁 SEND EXCEL FILE (.XLSX)\n\n"
        "RULES:\n"
        "• ID CAN BE IN ANY COLUMN\n"
        "• ONLY ID → ✅ VALID\n"
        "• EXTRA TEXT → ❌ INVALID"
    )

# ================= CONTINUE =================
@bot.callback_query_handler(func=lambda c: c.data == "continue")
def cont(c):
    if not is_joined(c.from_user.id):
        join_msg(c.message.chat.id)
        return
    start(c.message)

# ================= EXCEL RECEIVE =================
@bot.message_handler(content_types=["document"])
@private_only
def excel_receive(msg):
    uid = msg.chat.id
    all_users.add(uid)

    if session.get(uid, {}).get("state") != "WAIT_EXCEL":
        return

    file = bot.get_file(msg.document.file_id)
    data = bot.download_file(file.file_path)

    # ✅ FIX: SAVE FILE IN CURRENT DIRECTORY (NO /tmp)
    path = f"./{uid}.xlsx"
    with open(path, "wb") as f:
        f.write(data)

    wb = load_workbook(path, read_only=True)
    sheet = wb.active

    data_map = {}
    dup = set()

    for row in sheet.iter_rows(values_only=True):
        cells = [str(c).strip() for c in row if c is not None]
        ids = [c for c in cells if c.isdigit()]
        if not ids:
            continue

        i = ids[0]
        if i in data_map:
            dup.add(i)
            continue

        invalid = any(c != i for c in cells)
        data_map[i] = "INVALID" if invalid else "VALID"

    session[uid] = {"state": "WAIT_IDS", "data": data_map, "dup": dup}
    bot.send_message(uid, "📋 SEND ID LIST (ONE PER LINE)")

# ================= ID CHECK =================
@bot.message_handler(func=lambda m: session.get(m.chat.id, {}).get("state") == "WAIT_IDS")
@private_only
def id_check(msg):
    uid = msg.chat.id
    ids = re.findall(r"\d+", msg.text)

    valid, invalid, dup, nf = [], [], [], []
    seen = set()

    for i in ids:
        if i in seen:
            dup.append(i)
            continue
        seen.add(i)

        if i in session[uid]["dup"]:
            dup.append(i)
        elif i not in session[uid]["data"]:
            nf.append(i)
        elif session[uid]["data"][i] == "VALID":
            valid.append(i)
        else:
            invalid.append(i)

    def block(title, arr):
        return f"{title} ({len(arr)})\n" + ("\n".join(arr) if arr else "NONE")

    bot.send_message(
        uid,
        "✅ RESULT\n\n"
        f"{block('VALID', valid)}\n\n"
        f"{block('INVALID', invalid)}\n\n"
        f"{block('DUPLICATE', dup)}\n\n"
        f"{block('NOT FOUND', nf)}"
    )

    session.pop(uid, None)
    try:
        os.remove(f"./{uid}.xlsx")
    except:
        pass

    bot.send_message(
        uid,
        "✅ JOB COMPLETED\n\n"
        "ALL SESSION DATA HAS BEEN CLEARED.\n\n"
        "CLICK /start TO CHECK AGAIN.\n\n"
        "UTILITY COMMANDS:\n"
        "• /count – COUNT IDS\n"
        "• /aura – MERGE IDS\n"
        "• /aura2 – SPLIT IDS"
    )

    send_thanks(uid)

# ================= COUNT =================
@bot.message_handler(commands=["count"])
@private_only
def count(msg):
    utility[msg.chat.id] = {"mode": "COUNT"}
    bot.send_message(msg.chat.id, "📊 SEND IDS IN ONE MESSAGE")

@bot.message_handler(func=lambda m: utility.get(m.chat.id, {}).get("mode") == "COUNT")
@private_only
def do_count(msg):
    ids = re.findall(r"\d+", msg.text)
    bot.send_message(msg.chat.id, f"📊 TOTAL IDS: {len(ids)}")
    utility.pop(msg.chat.id)
    send_thanks(msg.chat.id)

# ================= AURA =================
@bot.message_handler(commands=["aura"])
@private_only
def aura(msg):
    utility[msg.chat.id] = {"mode": "AURA", "data": []}
    bot.send_message(
        msg.chat.id,
        "📥 SEND IDS / FORWARDED MESSAGES\nKEEP SENDING…\nSEND /done WHEN FINISHED"
    )

# ================= AURA2 =================
@bot.message_handler(commands=["aura2"])
@private_only
def aura2(msg):
    utility[msg.chat.id] = {"mode": "AURA2", "data": []}
    bot.send_message(
        msg.chat.id,
        "📥 SEND BIG TEXT\nKEEP SENDING…\nSEND /done WHEN FINISHED"
    )

# ================= COLLECT =================
@bot.message_handler(func=lambda m: m.chat.id in utility and m.text != "/done")
@private_only
def collect(msg):
    utility[msg.chat.id]["data"].extend(re.findall(r"\d+", msg.text))

# ================= DONE =================
@bot.message_handler(commands=["done"])
@private_only
def done(msg):
    u = utility.get(msg.chat.id)
    if not u:
        return

    out = "\n".join([f"{i+1}. {v}" for i, v in enumerate(u["data"])])
    bot.send_message(msg.chat.id, f"✅ RESULT\n\n{out}")
    utility.pop(msg.chat.id)
    send_thanks(msg.chat.id)

# ================= BROADCAST =================
@bot.message_handler(commands=["broadcast"])
@private_only
def broadcast(msg):
    if msg.from_user.id != OWNER_ID:
        return

    bot.send_message(msg.chat.id, "📢 SEND MESSAGE TO BROADCAST")
    bot.register_next_step_handler(msg, do_broadcast)

def do_broadcast(msg):
    sent = 0
    for uid in list(all_users):
        try:
            bot.send_message(uid, msg.text)
            sent += 1
        except:
            continue
    bot.send_message(msg.chat.id, f"✅ BROADCAST SENT TO {sent} USERS")

# ================= RUN =================
bot.infinity_polling()
